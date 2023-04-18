import requests
from bs4 import BeautifulSoup
import openpyxl
import json

with open('data.json') as f:
    data = json.load(f)

urls = []
quantities = []
product_names = []

for product, values in data.items():
    urls.append(values['url'])
    quantities.append(values['quantity'])
    product_names.append(product)

all_prices = {}
wb = openpyxl.Workbook()
ws = wb.active

# Write excel headers
ws.cell(row=1, column=1).value = "Shop"
column = 2
for product_name in product_names:
    ws.cell(row=1, column=column).value = product_name
    column += 1
ws.cell(row=1, column=column).value = "Sum"


# Web scraping
for i in range(len(urls)):
    url = urls[i]
    quantity = quantities[i]
    response = requests.get(url)
    soup = BeautifulSoup(response.text, 'html.parser')
    divs = soup.find_all(attrs={'data-price': True, 'data-shopurl': True})
    for div in divs:
        price = float(div.get('data-price'))
        shop = div.get('data-shopurl')
        if shop not in all_prices.keys():
            all_prices[shop] = {}
            all_prices[shop]['prices'] = [0] * len(urls)
            all_prices[shop]['count'] = 0
        all_prices[shop]['prices'][i] = price * quantity
        all_prices[shop]['count'] += 1

# Write shop names and prices in the remaining rows and columns
row = 2
for shop, data in all_prices.items():
    count = data['count']
    prices = data['prices']
    ws.cell(row=row, column=1).value = shop
    for j in range(len(prices)):
        if prices[j] != 0:
            ws.cell(row=row, column=j+2).value = prices[j]
            ws.cell(row=row, column=j+2).number_format = "#,##0.00 zł"
    ws.cell(row=row, column=column).value = sum(prices)
    ws.cell(row=row, column=column).number_format = "#,##0.00 zł"
    row += 1

# Set the table style
table = openpyxl.worksheet.table.Table(displayName="Table1", ref="A1:{}{}".format(openpyxl.utils.get_column_letter(column), row-1))
style = openpyxl.worksheet.table.TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
table.tableStyleInfo = style
ws.add_table(table)

# Save to the Excel file
wb.save("basket.xlsx")